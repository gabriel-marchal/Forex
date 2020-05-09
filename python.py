import json
from openpyxl import Workbook, load_workbook
from datetime import datetime , timedelta, date
import time
import os

import pyodbc

# 'Driver={ODBC Driver 17 for SQL Server};'
try:
    server = 'forexbd.database.windows.net'
    database = 'Forex'
    username = 'gabrielmarchal206'
    password = 'MS1!murkdeep'
    driver= '{ODBC Driver 17 for SQL Server}'
    cnxn = pyodbc.connect(DRIVER=driver,SERVER=server,PORT=1433,DATABASE=database,UID=username,PWD=password)
    cursor = cnxn.cursor()
except:
    print("Could not connect to SQL server")

wb = Workbook()
ws = wb.active

def insertSQL(jsonfile):
    print("Executing INSERT at "+ jsonfile["Date"])
    cursor.execute("""INSERT INTO Exchanges (Date,AUDCAD,AUDJPY,AUDNZD,AUDUSD,AUDCHF,CADJPY,CADCHF,CHFJPY,
                    EURAUD,EURCAD,EURGBP,EURNZD,EURUSD,EURCHF,EURJPY,GBPAUD,GBPCAD,
                    GBPJPY,GBPNZD, GBPUSD,GBPCHF,NZDCAD,NZDJPY,NZDUSD,USDCAD,USDCHF,USDJPY) 
                    values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", 
                    (jsonfile["Date"],jsonfile["AUDCAD"], jsonfile["AUDJPY"], jsonfile["AUDNZD"], jsonfile["AUDUSD"], 
                    jsonfile["AUDCHF"], jsonfile["CADJPY"], jsonfile["CADCHF"],jsonfile["CHFJPY"], 
                    jsonfile["EURAUD"], jsonfile["EURCAD"], jsonfile["EURGBP"], jsonfile["EURNZD"], jsonfile["EURUSD"], jsonfile["EURCHF"],
                    jsonfile["EURJPY"], jsonfile["GBPAUD"], jsonfile["GBPCAD"], jsonfile["GBPJPY"], jsonfile["GBPNZD"],
                    jsonfile["GBPUSD"], jsonfile["GBPCHF"], jsonfile["NZDCAD"], jsonfile["NZDJPY"], jsonfile["NZDUSD"],
                    jsonfile["USDCAD"], jsonfile["USDCHF"], jsonfile["USDJPY"]) )
    cursor.commit()
    print("Commit completed")
    
  

def convertdate(runDate):
    
    year= str(runDate.year)
    month = str(runDate.month)
    day = str(runDate.day)

    if len(month) == 1:
        month = "0"+month    
    if len(day) == 1:
        day = "0"+day
        
    date = year +"-"+month+"-"+day
    return date
  
def inputworksheet(jsonfile, i):
    #Input Column A jsonfile["Date"] , column B jsonfile["AUDCAD"] ....
    i = str(i+1)

    Acolumn = str("A"+i)
    Bcolumn = str("B"+i)
    Ccolumn = str("C"+i)
    Dcolumn = str("D"+i)
    Ecolumn = str("E"+i)
    Fcolumn = str("F"+i)
    Gcolumn = str("G"+i)
    Hcolumn = str("H"+i)
    Icolumn = str("I"+i)
    Jcolumn = str("J"+i)
    Kcolumn = str("K"+i)
    Lcolumn = str("L"+i)
    Mcolumn = str("M"+i)
    Ncolumn = str("N"+i)
    Ocolumn = str("O"+i)
    Pcolumn = str("P"+i)
    Qcolumn = str("Q"+i)
    Rcolumn = str("R"+i)
    Scolumn = str("S"+i)
    Tcolumn = str("T"+i)
    Ucolumn = str("U"+i)
    Vcolumn = str("V"+i)
    Wcolumn = str("W"+i)
    Xcolumn = str("X"+i)
    Ycolumn = str("Y"+i)
    Zcolumn = str("Z"+i)
    AAcolumn = str("AA"+i)
    ABcolumn = str("AB"+i)


    ws[Acolumn]= jsonfile["Date"]
    ws[Bcolumn] = jsonfile["AUDCAD"]
    ws[Ccolumn] = jsonfile["AUDJPY"]
    ws[Dcolumn] = jsonfile["AUDNZD"]
    ws[Ecolumn] = jsonfile["AUDUSD"]
    ws[Fcolumn] = jsonfile["AUDCHF"]
    ws[Gcolumn] = jsonfile["CADJPY"]
    ws[Hcolumn] = jsonfile["CADCHF"]
    ws[Icolumn] = jsonfile["CHFJPY"]
    ws[Jcolumn] = jsonfile["EURAUD"]
    ws[Kcolumn] = jsonfile["EURCAD"]
    ws[Lcolumn] = jsonfile["EURGBP"]
    ws[Mcolumn] = jsonfile["EURNZD"]
    ws[Ncolumn] = jsonfile["EURUSD"]
    ws[Ocolumn] = jsonfile["EURCHF"]
    ws[Pcolumn] = jsonfile["EURJPY"]
    ws[Qcolumn] = jsonfile["GBPAUD"]
    ws[Rcolumn] = jsonfile["GBPCAD"]
    ws[Scolumn] = jsonfile["GBPJPY"]
    ws[Tcolumn] = jsonfile["GBPNZD"]
    ws[Ucolumn] = jsonfile["GBPUSD"]
    ws[Vcolumn] = jsonfile["GBPCHF"]
    ws[Wcolumn] = jsonfile["NZDCAD"]
    ws[Xcolumn] = jsonfile["NZDJPY"]
    ws[Ycolumn] = jsonfile["NZDUSD"]
    ws[Zcolumn] = jsonfile["USDCAD"]
    ws[AAcolumn] = jsonfile["USDCHF"]
    ws[ABcolumn] = jsonfile["USDJPY"]

    wb.save(filename = "excelpairs.xlsx")
    print("Excel save with " + jsonfile["Date"])
    

def openfile(date, i):
    try:
        f = open("currencypairs"+date+".txt", "r")
        content = f.read()
        print(content)
        f.close()
        os.remove("currencypairs"+date+".txt")
        jsonfile = json.loads(content)
        print("File loaded " + jsonfile["Date"])
        inputworksheet(jsonfile, i)
        insertSQL(jsonfile)
    except:
        print("No file found for "+ date +". Enter sleep "+ str(datetime.now()))
        time.sleep(10)
        openfile(date, i)

def main():
    ws["A1"] = "Date / Pair"
    ws["B1"] = "AUDCAD"
    ws["C1"] = "AUDJPY"
    ws["D1"] = "AUDNZD"
    ws["E1"] = "AUDUSD"
    ws["F1"] = "AUDCHF"
    ws["G1"] = "CADJPY"
    ws["H1"] = "CADCHF"
    ws["I1"] = "CHFJPY"
    ws["J1"] = "EURAUD"
    ws["K1"] = "EURCAD"
    ws["L1"] = "EURGBP"
    ws["M1"] = "EURNZD"
    ws["N1"] = "EURUSD"
    ws["O1"] = "EURCHF"
    ws["P1"] = "EURJPY"
    ws["Q1"] = "GBPAUD"
    ws["R1"] = "GBPCAD"
    ws["S1"] = "GBPJPY"
    ws["T1"] = "GBPNZD"
    ws["U1"] = "GBPUSD"
    ws["V1"] = "GBPCHF"
    ws["W1"] = "NZDCAD"
    ws["X1"] = "NZDJPY"
    ws["Y1"] = "NZDUSD"
    ws["Z1"] = "USDCAD"
    ws["AA1"] = "USDCHF"
    ws["AB1"] = "USDJPY"
    i=1
    while True:
        # date = datetime.now()  - timedelta(days = i)
        date = datetime(2013,12,12)  - timedelta(days = i)
        newdate = convertdate(date)
        print("########################################## \nRunning date " + newdate)
        openfile(newdate, i)
        i=i+1
        

main()

